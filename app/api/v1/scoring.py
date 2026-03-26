"""
Scoring endpoints.
"""
from typing import List
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import io

from app.dependencies import get_db
from app.core.scoring.score_engine import ScoreEngine
from app.schemas.scoring import (
    ScoringRunCreate, ScoringRunResponse, ScoringRunStatus,
    ScoringResultsResponse, KeywordScoreResponse
)
from app.database.models import ScoringRun, KeywordScore, Keyword


router = APIRouter()


@router.post("/runs", response_model=ScoringRunResponse, status_code=201)
def create_scoring_run(
    run_data: ScoringRunCreate,
    db: Session = Depends(get_db)
):
    """
    Create a new scoring run.
    This initializes a scoring session with specified capacities.
    """
    engine = ScoreEngine(db)
    scoring_run = engine.create_scoring_run(
        ads_capacity=run_data.ads_capacity,
        seo_capacity=run_data.seo_capacity,
        social_capacity=run_data.social_capacity,
        default_relevance_coefficient=run_data.default_relevance_coefficient,
        run_name=run_data.run_name,
        company_url=run_data.company_url,
        competitor_urls=run_data.competitor_urls
    )
    return ScoringRunResponse.model_validate(scoring_run)


@router.post("/runs/{run_id}/execute", response_model=dict)
def execute_scoring(
    run_id: int,
    db: Session = Depends(get_db)
):
    """
    Execute scoring for all keywords in the run.
    Calculates ADS, SEO, and SOCIAL scores for all active keywords.
    """
    engine = ScoreEngine(db)
    
    try:
        result = engine.run_scoring(run_id)
        return result
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.get("/runs", response_model=List[ScoringRunStatus])
def list_scoring_runs(
    skip: int = 0,
    limit: int = 20,
    db: Session = Depends(get_db)
):
    """List all scoring runs."""
    runs = db.query(ScoringRun).order_by(ScoringRun.created_at.desc()).offset(skip).limit(limit).all()
    return [ScoringRunStatus.model_validate(run) for run in runs]


@router.get("/runs/{run_id}", response_model=ScoringRunStatus)
def get_scoring_run(run_id: int, db: Session = Depends(get_db)):
    """Get details of a specific scoring run."""
    run = db.query(ScoringRun).filter(ScoringRun.id == run_id).first()
    if not run:
        raise HTTPException(status_code=404, detail="Scoring run not found")
    return ScoringRunStatus.model_validate(run)


@router.delete("/runs/{run_id}", status_code=204)
def delete_scoring_run(run_id: int, db: Session = Depends(get_db)):
    """Delete a scoring run."""
    from app.database import crud
    success = crud.delete_scoring_run(db, run_id)
    if not success:
        raise HTTPException(status_code=404, detail="Scoring run not found")
    return None


@router.get("/runs/{run_id}/scores", response_model=ScoringResultsResponse)
def get_scoring_results(
    run_id: int,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    """Get scored keywords for a run."""
    run = db.query(ScoringRun).filter(ScoringRun.id == run_id).first()
    if not run:
        raise HTTPException(status_code=404, detail="Scoring run not found")
    
    scores = (
        db.query(KeywordScore, Keyword)
        .join(Keyword, KeywordScore.keyword_id == Keyword.id)
        .filter(KeywordScore.scoring_run_id == run_id)
        .limit(limit)
        .all()
    )
    
    return ScoringResultsResponse(
        scoring_run_id=run_id,
        status=run.status,
        total_scored=len(scores),
        scores=[
            KeywordScoreResponse(
                keyword_id=score.keyword_id,
                keyword=keyword.keyword,
                ads_score=score.ads_score,
                seo_score=score.seo_score,
                social_score=score.social_score,
                ads_rank=score.ads_rank,
                seo_rank=score.seo_rank,
                social_rank=score.social_rank
            )
            for score, keyword in scores
        ]
    )


@router.get("/runs/{run_id}/top/{channel}")
def get_top_by_channel(
    run_id: int,
    channel: str,
    limit: int = 10,
    db: Session = Depends(get_db)
):
    """Get top scoring keywords for a specific channel."""
    if channel.upper() not in ['ADS', 'SEO', 'SOCIAL']:
        raise HTTPException(status_code=400, detail="Invalid channel. Use ADS, SEO, or SOCIAL")
    
    engine = ScoreEngine(db)
    top_keywords = engine.get_top_keywords_by_channel(run_id, channel.upper(), limit)
    
    return {
        "channel": channel.upper(),
        "top_keywords": top_keywords
    }


@router.get("/runs/{run_id}/export/xlsx")
def export_scoring_xlsx(
    run_id: int,
    db: Session = Depends(get_db)
):
    """Export scoring results as XLSX file (native Excel format, no date auto-format issues)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, numbers

    run = db.query(ScoringRun).filter(ScoringRun.id == run_id).first()
    if not run:
        raise HTTPException(status_code=404, detail="Scoring run not found")
    
    scores = (
        db.query(KeywordScore, Keyword)
        .join(Keyword, KeywordScore.keyword_id == Keyword.id)
        .filter(KeywordScore.scoring_run_id == run_id)
        .all()
    )
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Skorlama Sonuçları"
    
    # Header row
    headers = [
        'Keyword ID', 'Keyword', 'Sektör',
        'Aylık Hacim', 'Trend 3M (%)', 'Trend 12M (%)', 'Rekabet Skoru',
        'ADS Skor', 'ADS Sıra', 'SEO Skor', 'SEO Sıra',
        'SOCIAL Skor', 'SOCIAL Sıra'
    ]
    ws.append(headers)
    
    # Style header row — bold + auto-filter
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font
    ws.auto_filter.ref = ws.dimensions
    
    # Data rows
    for score, keyword in scores:
        ws.append([
            keyword.id,
            keyword.keyword,
            keyword.sector or '',
            keyword.monthly_volume or 0,
            float(keyword.trend_3m) if keyword.trend_3m is not None else 0,
            float(keyword.trend_12m) if keyword.trend_12m is not None else 0,
            float(keyword.competition_score) if keyword.competition_score is not None else 0,
            float(score.ads_score) if score.ads_score else 0,
            score.ads_rank or 0,
            float(score.seo_score) if score.seo_score else 0,
            score.seo_rank or 0,
            float(score.social_score) if score.social_score else 0,
            score.social_rank or 0
        ])
    
    # Set number format on score columns to prevent date interpretation
    # Columns: E(5)=Trend3M, F(6)=Trend12M, G(7)=Rekabet,
    #          H(8)=ADS Skor, J(10)=SEO Skor, L(12)=SOCIAL Skor
    score_cols = [5, 6, 7, 8, 10, 12]  # 1-indexed column numbers
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in score_cols:
            cell = row[col_idx - 1]  # 0-indexed in row tuple
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.0000'
    
    # Auto-fit column widths
    col_widths = [12, 40, 8, 12, 12, 13, 13, 12, 9, 12, 9, 13, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    
    # Write to bytes buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"scoring_run_{run_id}_{run.run_name or 'export'}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
