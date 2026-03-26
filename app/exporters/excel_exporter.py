"""
Excel Exporter Module.

Excel dosyası olarak export eder.
13 sheet ile kapsamlı veri çıktısı.
Türkçe karakter desteği (openpyxl UTF-8).
"""
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.exporters.base_exporter import BaseExporter
from app.schemas.export import ExportSectionEnum, FullReport, BrandProfileData


class ExcelExporter(BaseExporter):
    """
    Excel dosyası olarak export eder.
    
    13 Sheet:
    1. Özet
    2. Tüm Kelimeler
    3-5. ADS/SEO/SOCIAL Havuzları
    6. Stratejik
    7. SEO İçerikler
    8-11. Ads (Gruplar, Başlıklar, Açıklamalar, Negatifler)
    12-13. Sosyal (Kategoriler/Fikirler, İçerikler)
    """
    
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def export(
        self,
        scoring_run_id: int,
        sections: List[ExportSectionEnum],
        filepath: str
    ) -> str:
        """Export işlemini yapar."""
        data = self.collect_data(scoring_run_id, sections)
        
        wb = Workbook()
        
        # İlk default sheet'i kullan
        ws = wb.active
        ws.title = "Özet"
        self._add_summary_sheet(ws, data)

        if self._should_include(ExportSectionEnum.BRAND_PROFILE, sections):
            self._add_brand_profile_sheets(wb, data.brand_profile)
        
        # Diğer sheetler
        if self._should_include(ExportSectionEnum.SCORING, sections) and data.scoring:
            self._add_keywords_sheet(wb, data.scoring)
        
        if self._should_include(ExportSectionEnum.CHANNELS, sections) and data.channels:
            self._add_channel_sheets(wb, data.channels)
        
        if self._should_include(ExportSectionEnum.SEO_CONTENT, sections) and data.seo_contents:
            self._add_seo_content_sheet(wb, data.seo_contents)
        
        if self._should_include(ExportSectionEnum.ADS, sections) and data.ads:
            self._add_ads_sheets(wb, data.ads)
        
        if self._should_include(ExportSectionEnum.SOCIAL, sections) and data.social:
            self._add_social_sheets(wb, data.social)
        
        wb.save(filepath)
        return filepath
    
    def _style_header(self, ws, row: int, cols: int):
        """Header satırını stillendirir."""
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
    
    def _add_summary_sheet(self, ws, data: FullReport):
        """Özet sheet'i."""
        ws.append(['DIGITUS ENGINE RAPORU'])
        ws.append([])
        ws.append(['Metrik', 'Değer'])
        self._style_header(ws, 3, 2)
        
        if data.summary:
            ws.append(['Scoring Run ID', data.scoring_run_id])
            ws.append(['Tarih', data.generated_at.strftime('%Y-%m-%d %H:%M')])
            ws.append(['Toplam Kelime', data.summary.total_keywords])
            ws.append([])
            ws.append(['Kanal Dağılımı', ''])
            ws.append(['ADS Havuzu', data.summary.ads_count])
            ws.append(['SEO Havuzu', data.summary.seo_count])
            ws.append(['SOCIAL Havuzu', data.summary.social_count])
            ws.append(['Stratejik', data.summary.strategic_count])
            ws.append([])
            ws.append(['Üretilen İçerik', ''])
            ws.append(['SEO+GEO İçerik', data.summary.seo_content_count])
            ws.append(['Reklam Grubu', data.summary.ad_group_count])
            ws.append(['Sosyal İçerik', data.summary.social_content_count])
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _add_keywords_sheet(self, wb: Workbook, scoring):
        """Tüm kelimeler sheet'i."""
        ws = wb.create_sheet("Tüm Kelimeler")
        
        headers = ['Kelime', 'Hacim', 'Trend 3ay', 'Trend 12ay', 'Rekabet',
                   'ADS Skor', 'SEO Skor', 'SOCIAL Skor',
                   'ADS Rank', 'SEO Rank', 'SOCIAL Rank', 'Birincil', 'Niyet']
        ws.append(headers)
        self._style_header(ws, 1, len(headers))
        
        for kw in scoring.keywords:
            ws.append([
                kw.keyword,
                kw.volume,
                kw.trend_3m,
                kw.trend_12m,
                kw.competition,
                kw.ads_score,
                kw.seo_score,
                kw.social_score,
                kw.ads_rank,
                kw.seo_rank,
                kw.social_rank,
                kw.primary_channel,
                kw.intent
            ])
        
        # Kolon genişlikleri
        ws.column_dimensions['A'].width = 35
        for col in 'BCDEFGHIJKLM':
            ws.column_dimensions[col].width = 12

    def _add_brand_profile_sheets(self, wb: Workbook, brand_profile: Optional[BrandProfileData]):
        """Marka profili ve rakip bilgisi sheetleri."""
        ws = wb.create_sheet("Marka Profili")
        ws.append(['Alan', 'Deger'])
        self._style_header(ws, 1, 2)

        if not brand_profile:
            ws.append(['Durum', 'Profil bulunamadi'])
            ws.column_dimensions['A'].width = 28
            ws.column_dimensions['B'].width = 80
            return

        ws.append(['Durum', brand_profile.status])
        ws.append(['Firma URL', brand_profile.company_url or ''])
        ws.append(['Firma Adi', brand_profile.company_name or ''])
        ws.append(['Sektor', brand_profile.sector or ''])
        ws.append(['Hedef Kitle', brand_profile.target_audience or ''])
        ws.append(['Urunler', ', '.join(brand_profile.products)])
        ws.append(['Hizmetler', ', '.join(brand_profile.services)])
        ws.append(['Kullanim Alanlari', ', '.join(brand_profile.use_cases)])
        ws.append(['Cozulen Problemler', ', '.join(brand_profile.problems_solved)])
        ws.append(['Marka Terimleri', ', '.join(brand_profile.brand_terms)])
        ws.append(['Dislanan Temalar', ', '.join(brand_profile.exclude_themes)])
        ws.append(['Anchor Textler', ' | '.join(brand_profile.anchor_texts)])
        ws.append(['Hata Mesaji', brand_profile.error_message or ''])

        if brand_profile.validation_warnings:
            ws.append(['Dogrulama Uyarilari', ' | '.join(brand_profile.validation_warnings)])

        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 100

        src = wb.create_sheet("Profil Kaynaklar")
        src.append(['URL', 'Baslik', 'Durum'])
        self._style_header(src, 1, 3)
        if brand_profile.source_pages:
            for page in brand_profile.source_pages:
                if not isinstance(page, dict):
                    continue
                src.append([page.get('url', ''), page.get('title', ''), page.get('status', '')])
        else:
            src.append(['', 'Kaynak sayfa kaydi yok', ''])
        src.column_dimensions['A'].width = 60
        src.column_dimensions['B'].width = 50
        src.column_dimensions['C'].width = 12

        comp = wb.create_sheet("Rakipler")
        comp.append(['URL', 'Durum', 'Skor', 'Ozet'])
        self._style_header(comp, 1, 4)
        if brand_profile.competitors:
            for c in brand_profile.competitors:
                comp.append([
                    c.url or '',
                    c.status or '',
                    c.consistency_score,
                    c.summary or ''
                ])
        elif brand_profile.competitor_urls:
            for url in brand_profile.competitor_urls:
                comp.append([url, 'analiz_yok', None, ''])
        else:
            comp.append(['', 'Rakip kaydi yok', None, ''])
        comp.column_dimensions['A'].width = 55
        comp.column_dimensions['B'].width = 20
        comp.column_dimensions['C'].width = 12
        comp.column_dimensions['D'].width = 80
    
    def _add_channel_sheets(self, wb: Workbook, channels):
        """Kanal havuzları sheetleri."""
        for name, pool in [('ADS Havuzu', channels.ads), 
                          ('SEO Havuzu', channels.seo), 
                          ('SOCIAL Havuzu', channels.social)]:
            ws = wb.create_sheet(name)
            
            headers = ['Sıra', 'Kelime', 'Skor', 'Vektör Yakınlığı', 'Çarpım Skoru', 'Niyet']
            ws.append(headers)
            self._style_header(ws, 1, len(headers))
            
            for i, kw in enumerate(pool.keywords, 1):
                score = kw.ads_score or kw.seo_score or kw.social_score or 0
                ws.append([
                    i,
                    kw.keyword,
                    score,
                    kw.vector_similarity,
                    kw.vector_adjusted_score,
                    kw.intent,
                ])
            
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 16
            ws.column_dimensions['E'].width = 14
            ws.column_dimensions['F'].width = 20
        
        # Stratejik
        ws = wb.create_sheet("Stratejik")
        headers = ['Kelime', 'ADS Rank', 'SEO Rank', 'ADS Skor', 'SEO Skor']
        ws.append(headers)
        self._style_header(ws, 1, len(headers))
        
        for kw in channels.strategic:
            ws.append([kw.keyword, kw.ads_rank, kw.seo_rank, kw.ads_score, kw.seo_score])
        
        ws.column_dimensions['A'].width = 40
    
    def _add_seo_content_sheet(self, wb: Workbook, seo_contents):
        """SEO içerikler sheet'i."""
        ws = wb.create_sheet("SEO İçerikler")
        
        headers = ['Kelime', 'Başlık', 'URL', 'Kelime Sayısı', 'SEO Skor', 'GEO Skor', 'Combined']
        ws.append(headers)
        self._style_header(ws, 1, len(headers))
        
        for c in seo_contents.contents:
            ws.append([
                c.keyword,
                c.title,
                c.url_suggestion,
                c.word_count,
                c.seo_score,
                c.geo_score,
                c.combined_score
            ])
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 30
    
    def _add_ads_sheets(self, wb: Workbook, ads):
        """Ads sheetleri (4 tane)."""
        # Reklam Grupları
        ws = wb.create_sheet("Reklam Grupları")
        headers = ['Grup Adı', 'Hedef Kelimeler', 'Başlık Sayısı', 'Açıklama Sayısı', 'Negatif Sayısı']
        ws.append(headers)
        self._style_header(ws, 1, len(headers))
        
        for g in ads.ad_groups:
            ws.append([
                g.group_name,
                ', '.join(g.target_keywords[:3]),
                len(g.headlines),
                len(g.descriptions),
                len(g.negative_keywords)
            ])
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
        
        # Başlıklar
        ws = wb.create_sheet("Başlıklar")
        headers = ['Grup', 'Başlık', 'Tip', 'DKI']
        ws.append(headers)
        self._style_header(ws, 1, len(headers))
        
        for g in ads.ad_groups:
            for h in g.headlines:
                ws.append([g.group_name, h.headline_text, h.headline_type, '✔' if h.is_dki else ''])
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35
        
        # Açıklamalar
        ws = wb.create_sheet("Açıklamalar")
        headers = ['Grup', 'Açıklama', 'Tip']
        ws.append(headers)
        self._style_header(ws, 1, len(headers))
        
        for g in ads.ad_groups:
            for d in g.descriptions:
                ws.append([g.group_name, d.description_text, d.description_type])
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 80
        
        # Negatifler
        ws = wb.create_sheet("Negatifler")
        headers = ['Grup', 'Kelime', 'Eşleme', 'Sebep']
        ws.append(headers)
        self._style_header(ws, 1, len(headers))
        
        for g in ads.ad_groups:
            for n in g.negative_keywords:
                ws.append([g.group_name, n.keyword, n.match_type, n.reason])
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['D'].width = 40
    
    def _add_social_sheets(self, wb: Workbook, social):
        """Sosyal medya sheetleri."""
        # Kategoriler & Fikirler
        ws = wb.create_sheet("Sosyal Fikirler")
        headers = ['Başlık', 'Platform', 'Format', 'Trend Alignment', 'Seçildi']
        ws.append(headers)
        self._style_header(ws, 1, len(headers))
        
        for i in social.ideas:
            ws.append([
                i.get('title', ''),
                i.get('platform', ''),
                i.get('format', ''),
                i.get('trend_alignment', 0),
                '✔' if i.get('is_selected') else ''
            ])
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        
        # İçerikler
        ws = wb.create_sheet("Sosyal İçerikler")
        headers = ['Fikir', 'Platform', 'Hook 1', 'Caption', 'Hashtagler']
        ws.append(headers)
        self._style_header(ws, 1, len(headers))
        
        for c in social.contents:
            hook1 = c.hooks[0].text if c.hooks else ''
            ws.append([
                c.idea_title,
                c.platform,
                hook1[:50],
                c.caption[:100] if c.caption else '',
                ' '.join([f'#{t}' for t in c.hashtags[:5]])
            ])
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 40

